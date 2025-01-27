import openpyxl
from django.shortcuts import render,redirect
from django.http import HttpResponse
import os
from django.conf import settings
from myapp.firebase import auth,upload_to_firebase
import openpyxl
import os
from django.conf import settings
from django.shortcuts import redirect

def update_column_headers(extracted_file_path, subject_code_map):
    # Load the extracted workbook
    wb = openpyxl.load_workbook(extracted_file_path)
    sheet = wb.active

    # Iterate through each column header in the first row
    for col in range(2, sheet.max_column + 1):  # Starting from column 2 to avoid "S.No"
        subject_code = sheet.cell(row=1, column=col).value
        # Check if the subject code exists in the subject_code_map
        if subject_code in subject_code_map:
            # Update the header with the mapped value
            sheet.cell(row=1, column=col, value=subject_code_map[subject_code])

    # Save the workbook with updated headers
    wb.save(extracted_file_path)

def process_subject_codes(sheet4):
    subject_code_map = {}
    for row in range(2, sheet4.max_row + 1):
        subject_code = sheet4.cell(row=row, column=2).value
        subject_name = sheet4.cell(row=row, column=3).value
        staff_name = sheet4.cell(row=row, column=4).value
        if subject_code:
            # Combine subject name and staff name with an underscore
            subject_code_map[subject_code] = f'{subject_code}/{subject_name}/{staff_name}'
    return subject_code_map

def home(request):
    download_url = None
    extracted_url = None
    email = None
    if request.method == 'POST' and 'excel_file1' in request.FILES and 'excel_file2' in request.FILES and 'excel_file3' in request.FILES and 'excel_file4' in request.FILES:
        department = request.POST.get('department')
        year = request.POST.get('year')
        study_year = request.POST.get('study_year')
        semester=request.POST.get('SEMESTER')
        section = request.POST.get('section')
        parent_key = f'{department}_{year}_{study_year}_{section}_{semester}'
        email = request.POST.get('email', None)
        excel_file1 = request.FILES['excel_file1']
        excel_file2 = request.FILES['excel_file2']
        excel_file3 = request.FILES['excel_file3']
        excel_file4 = request.FILES['excel_file4']
        wb1 = openpyxl.load_workbook(excel_file1)
        wb2 = openpyxl.load_workbook(excel_file2)
        wb3 = openpyxl.load_workbook(excel_file3)
        wb4 = openpyxl.load_workbook(excel_file4)
        sheet1 = wb1.active
        sheet2 = wb2.active
        sheet3 = wb3.active
        sheet4 = wb4.active
        subject_code_map = process_subject_codes(sheet4)
        print(subject_code_map)
        register_map = {}
        for row in range(2, sheet2.max_row + 1):
            reg_no = sheet2.cell(row=row, column=2).value
            sgpa = sheet2.cell(row=row, column=4).value
            cgpa = sheet2.cell(row=row, column=5).value
            if reg_no:
                register_map[reg_no] = (sgpa, cgpa)

        next_col_sgpa = sheet1.max_column + 1
        next_col_cgpa = next_col_sgpa + 1

        for row in range(2, sheet1.max_row + 1):
            reg_no = sheet1.cell(row=row, column=2).value
            if reg_no and reg_no in register_map:
                sgpa, cgpa = register_map[reg_no]
                sheet1.cell(row=row, column=next_col_sgpa, value=sgpa)
                sheet1.cell(row=row, column=next_col_cgpa, value=cgpa)

        merged_file_path = os.path.join(settings.MEDIA_ROOT, 'merged.xlsx')
        wb1.save(merged_file_path)
        download_url = f'{settings.MEDIA_URL}merged.xlsx'

        merged_wb = openpyxl.load_workbook(merged_file_path)
        merged_sheet = merged_wb.active

        extracted_wb = openpyxl.Workbook()
        extracted_sheet = extracted_wb.active

        headers = ["S.No"] + [merged_sheet.cell(row=2, column=col).value for col in range(2, merged_sheet.max_column + 1)]
        for col, header in enumerate(headers, start=1):
            extracted_sheet.cell(row=1, column=col, value=header)

        row_counter = 2
        for row in range(2, sheet3.max_row + 1):
            reg_no = sheet3.cell(row=row, column=2).value
            for merged_row in range(2, merged_sheet.max_row + 1):
                merged_reg_no = merged_sheet.cell(row=merged_row, column=2).value
                if reg_no == merged_reg_no:
                    extracted_sheet.cell(row=row_counter, column=1, value=row_counter - 1)  # S.No
                    for col in range(2, merged_sheet.max_column + 1):
                        extracted_sheet.cell(row=row_counter, column=col, value=merged_sheet.cell(row=merged_row, column=col).value)
                    row_counter += 1
                    break

        extracted_file_path = os.path.join(settings.MEDIA_ROOT, 'extracted.xlsx')
        extracted_wb.save(extracted_file_path)
        extracted_url = f'{settings.MEDIA_URL}extracted.xlsx'

        update_column_headers(extracted_file_path, subject_code_map)
        upload_to_firebase(email,parent_key, extracted_file_path)

    return render(request, 'home.html')

def copy_and_paste_columns(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    student_name_col = 3  # Assuming "student name" is in column 3
    sgpa_col = sheet.max_column - 1  # SGPA column, assuming second last column
    last_col = sheet.max_column + 1
    for row in range(2, sheet.max_row + 1):
        
        paste_col = last_col
        
        for col in range(student_name_col + 1, sgpa_col):
            value = sheet.cell(row=row-1, column=col).value
            sheet.cell(row=row-1, column=paste_col, value=value)
            paste_col += 1

    wb.save(file_path)

import pandas as pd
import json
def about(request):
    # Path to the extracted Excel file
    merged_file_path = os.path.join(settings.MEDIA_ROOT, 'extracted.xlsx')

    # Load the Excel file into a DataFrame
    df = pd.read_excel(merged_file_path)
    
    # Convert DataFrame to a list of dictionaries
    students = df.to_dict(orient='records')
    valid=make_valid_identifier(students)
    # Sort students by CGPA and SGPA
    sorted_students = sorted(valid, key=lambda x: (x['cgpa'], x['sgpa']), reverse=True)

    # Convert the sorted student data to JSON
    students_json = json.dumps(sorted_students)
    
    # Pass the JSON data to the template
    return render(request, 'about.html', {'students_json': students_json})

import re

def make_valid_identifier(data):
    # Function to convert dictionary keys to valid Python identifiers
    def convert_key(key):
        # Convert to lowercase and replace spaces with underscores
        key = key.strip().lower()
        # Remove any characters that are not alphanumeric or underscores
        key = re.sub(r'[^a-z0-9_]', '', key)
        return key
    
    # Iterate over the list of dictionaries and convert keys for each dictionary
    return [
        {convert_key(k): v for k, v in student.items()} 
        for student in data
    ]
from django.contrib import messages
from django.shortcuts import render, redirect
from myapp.firebase import auth
import firebase_admin
from firebase_admin import auth as admin_auth

# Ensure Firebase Admin SDK is initialized
if not firebase_admin._apps:
    firebase_admin.initialize_app()

def register(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect('register')

        try:
            # Create user
            user = auth.create_user(email=email, password=password)

            messages.success(request, f"Account created for {email}. Please log in.")
            return redirect('login')
        except firebase_admin._auth_utils.EmailAlreadyExistsError:
            messages.error(request, "Email already exists. Please use a different email.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    return render(request, 'register.html')

def login(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']

        try:
            user = auth.get_user_by_email(email)
            call()    
            # Here you should verify the password using Firebase SDK on the client-side
            messages.success(request, "Login successful.")
           
            return render(request, 'home.html', {'email': email})
        except firebase_admin._auth_utils.UserNotFoundError:
            messages.error(request, "Invalid credentials. Please check your email and password.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    return render(request, 'login.html')

from .firebase import get_report_data_by_email # Import the Firebase function
from django.http import JsonResponse
def fetch_report_data(request):
    if request.method == 'POST':
        print("Generate report function called")  # Log when the function is invoked

        # Your existing logic here
        return JsonResponse({'status': 'Report generation initiated'})

    return JsonResponse({'error': 'Invalid request method'}, status=400)
def call():
    c=get_report_data_by_email()
    print(c)

