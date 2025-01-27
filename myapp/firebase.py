import firebase_admin
from firebase_admin import credentials, auth
import os

# Setup paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIREBASE_CREDENTIALS_PATH = os.path.join(BASE_DIR, 'firebase', 'karthick.json')

# Initialize Firebase Admin SDK
if not firebase_admin._apps:
    cred = credentials.Certificate(FIREBASE_CREDENTIALS_PATH)
    firebase_admin.initialize_app(cred, {
            'databaseURL': 'https://login-eba02-default-rtdb.firebaseio.com/'
        })


# Export auth to be used in views.py
auth = auth
import firebase_admin
from firebase_admin import credentials, db
import openpyxl

def sanitize_key(key):
    # Replace invalid characters with underscores
    return key.replace('.', '_').replace('#', '_').replace('$', '_').replace('[', '_').replace(']', '_').replace('/', '_')

def get_user_ref(user_email):
    """ Check if the user exists, and return the reference to their node in the Firebase database """
    try:
        # Try to fetch the user based on their email (or any other unique identifier)
        user = auth.get_user_by_email(user_email)
        user_key = user.uid  # Firebase user ID (UID)
    except auth.UserNotFoundError:
        # User doesn't exist, you might want to create the user or handle it differently
        raise Exception("User not found. Please register first.")

    # Return the Firebase reference to the user's node
    return db.reference(sanitize_key(user_key))




def upload_to_firebase(user_email, parent_key, extracted_file_path):
    # Ensure Firebase is initialized

    # Get the reference to the user's node
    user_ref = get_user_ref(user_email)

    # Sanitize the parent key
    sanitized_parent_key = sanitize_key(parent_key)

    # Check if the parent key exists in the user's main node
    ref = user_ref.child(sanitized_parent_key)
    if ref.get() is not None:
        # Parent key exists directly in the user's node
        print(f"The parent key '{sanitized_parent_key}' already exists in the main node.")
        raise ValueError(f"The parent key '{sanitized_parent_key}' already exists. Confirm to overwrite.")

    # Check if the parent key exists in the 'databases' branch
    databases_ref = user_ref.child("databases")
    databases = databases_ref.get()

    # Convert `databases` to a dictionary if it's a list
    if isinstance(databases, list):
        # Convert list to a dictionary with indices as keys
        databases = {str(i): item for i, item in enumerate(databases) if item is not None}
    elif databases is None:
        # If `databases` does not exist, initialize it as an empty dictionary
        databases = {}

    # Check if the sanitized parent key exists in the `databases` branch
    for db_key, db_value in databases.items():
        if isinstance(db_value, dict) and db_value.get('name') == sanitized_parent_key:
            print(f"The parent key '{sanitized_parent_key}' already exists in the 'databases' branch.")
            raise ValueError(f"The parent key '{sanitized_parent_key}' already exists in 'databases'. Confirm to overwrite.")

    # Assign the next sequential key for the databases branch
    next_key = str(len(databases))

    # Insert the parent key into the 'databases' branch
    databases_ref.child(next_key).set({
        "name": sanitized_parent_key,
    })

    # Load the extracted workbook
    wb = openpyxl.load_workbook(extracted_file_path)
    sheet = wb.active

    # Get headers from the first row
    headers = [sanitize_key(sheet.cell(row=1, column=col).value) for col in range(1, sheet.max_column + 1)]

    # Iterate through each row and upload data to Firebase
    for row in range(2, sheet.max_row + 1):
        student_data = {header: sheet.cell(row=row, column=col).value for col, header in enumerate(headers, start=1)}

        # Generate a sanitized student key
        student_key = sanitize_key(f'student_{row-1}')
        
        # Upload data to Firebase
        ref.child(student_key).set(student_data)

    print(f"Data uploaded successfully under parent key '{sanitized_parent_key}'.")

def get_report_data_by_email():
    """Fetch report data for a user by their email."""
    user_ref = get_user_ref('karthickfrank007@gmail.com')  # Assuming this function returns a reference to the user

    if user_ref is None:
        print({'error': 'User not found'})
        return  # Exit the function if the user is not found

    # Fetch the branches associated with the user
    branches= user_ref.get()  # This will get all branches under this user's UID

    if branches:
        print("Available Branches:")
        print(branches)
    else:
        print("No branches found for this user.")
        return  # Exit the function if no branches are found

    # Print the email and branches data
    print({'email': 'karthickfrank007@gmail.com', 'branches': branches})

    # Return the branches data
   